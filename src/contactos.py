"""Lectura y limpieza de contactos desde Excel. No sabe nada de WhatsApp."""

import re
from dataclasses import dataclass, field

# Motivos de descarte
VACIO = "vacio"
BASURA = "basura"
NO_ES_MOVIL_CO = "no_es_movil_co"
LARGO_INVALIDO = "largo_invalido"
NO_CUADRA_CON_EL_PAIS = "no_cuadra_con_el_pais"
DUPLICADO = "duplicado"

# E.164 permite hasta 15 digitos contando el indicativo de pais. Por abajo, un
# numero con indicativo baja de 10 digitos solo si esta truncado: los "cortos"
# que aparecen en las listas suelen ser fijos viejos con un '+' puesto delante.
LARGO_MINIMO_INTERNACIONAL = 10
LARGO_MAXIMO_INTERNACIONAL = 15

# Cuantos digitos tiene un numero completo -indicativo incluido- en los paises
# de la lista. Un '+' delante no convierte una cedula en telefono: en la campana
# real, cuatro numeros como "+1037578093" (una cedula colombiana con +1 delante)
# pasaron la validacion general de largo y Meta los rechazo con el error 131009.
#
# Solo estan los paises cuyo formato conocemos con certeza. Para el resto basta
# el rango de E.164: preferimos dejar pasar un numero dudoso antes que descartar
# uno bueno de un pais que no tenemos tabulado.
LARGOS_POR_INDICATIVO = {
    "1": {11},        # EE.UU. y Canada
    "57": {12},       # Colombia
    "56": {11},       # Chile
    "51": {11},       # Peru
    "52": {12, 13},   # Mexico: 12 desde 2019, 13 con el "1" antiguo
    "593": {12},      # Ecuador
    "506": {11},      # Costa Rica
    "503": {11},      # El Salvador
    "507": {11},      # Panama
}


def _cuadra_con_su_pais(digitos):
    """El largo coincide con lo que usa ese indicativo? True si no lo sabemos.

    Se prueban los indicativos de mas largo a mas corto para que "593" gane
    sobre "59" y "5".
    """
    for tamano in (3, 2, 1):
        esperados = LARGOS_POR_INDICATIVO.get(digitos[:tamano])
        if esperados is not None:
            return len(digitos) in esperados
    return True


def normalizar_telefono(valor):
    """Convierte una celda de Excel en un telefono E.164 sin '+'.

    Devuelve (telefono, motivo). Si telefono es None, motivo explica por que se
    descarto. Si telefono tiene valor, motivo es "".

    Dos caminos segun lo que traiga la celda:

    - **Con '+'**: ya trae indicativo de pais. Se respeta tal cual, venga de
      donde venga, y solo se valida el largo.
    - **Sin '+'**: se asume Colombia, que es lo que exportan los CRM y planillas
      locales sin formato.

    Acepta str o int: en un Excel la misma columna puede venir de las dos formas.
    """
    if valor is None:
        return None, VACIO

    texto = str(valor).strip()
    if not texto:
        return None, VACIO

    digitos = re.sub(r"\D", "", texto)
    if not digitos:
        return None, BASURA

    if texto.startswith("+"):
        if not (LARGO_MINIMO_INTERNACIONAL <= len(digitos) <= LARGO_MAXIMO_INTERNACIONAL):
            # Poner un '+' delante no convierte un fijo de 8 digitos en internacional.
            return None, LARGO_INVALIDO
        if not _cuadra_con_su_pais(digitos):
            return None, NO_CUADRA_CON_EL_PAIS
        return digitos, ""

    # Sin indicativo: largos imposibles para cualquier telefono.
    if len(digitos) < 7 or len(digitos) > 13:
        return None, BASURA

    # Movil colombiano sin indicativo: 10 digitos que empiezan en 3.
    if len(digitos) == 10 and digitos.startswith("3"):
        return "57" + digitos, ""

    # Movil colombiano con indicativo: 57 + 10 digitos que empiezan en 3.
    if len(digitos) == 12 and digitos.startswith("57") and digitos[2] == "3":
        return digitos, ""

    return None, NO_ES_MOVIL_CO


# Restos tipicos de un export mal hecho (pandas escribe "nan" cuando la celda
# estaba vacia). Si no los quitamos, el mensaje dice "Hola, Andrea nan.".
BASURA_EN_NOMBRES = {"nan", "none", "null", "na", "-", "--", "."}


def normalizar_nombre(valor, por_defecto="Hola"):
    """Devuelve el primer nombre, capitalizado y sin residuos de export.

    "ANGELA SANCHEZ"  -> "Angela"
    "Andrea nan"      -> "Andrea"
    ""                -> por_defecto
    """
    if valor is None:
        return por_defecto

    texto = str(valor).strip()
    if not texto:
        return por_defecto

    palabras = [p for p in texto.split() if p.lower() not in BASURA_EN_NOMBRES]
    if not palabras:
        return por_defecto

    primero = palabras[0]
    # Un "nombre" sin ninguna letra no sirve para saludar a nadie.
    if not any(c.isalpha() for c in primero):
        return por_defecto

    return primero.capitalize()


@dataclass(frozen=True)
class Contacto:
    """Un destinatario listo para enviar."""

    nombre: str                  # primer nombre capitalizado, ej "Ana"
    telefono: str                # E.164 sin '+', ej "573001234567"
    fila: int                    # numero de fila en el Excel, para reportes
    crudo: dict = field(default_factory=dict)   # la fila original por nombre de columna


@dataclass(frozen=True)
class Descarte:
    """Una fila que no se puede usar, y por que."""

    fila: int
    nombre_crudo: str
    telefono_crudo: str
    motivo: str


def leer_contactos(
    ruta,
    hoja=None,
    columna_nombre="Nombre",
    columna_telefono="Teléfono",
    nombre_por_defecto="Hola",
):
    """Lee un .xlsx y devuelve (validos, descartes).

    Nunca lanza excepcion por una fila mala: la manda a descartes con su motivo.
    Solo falla si el archivo no existe o si faltan las columnas pedidas.
    """
    from openpyxl import load_workbook

    libro = load_workbook(ruta, read_only=True, data_only=True)
    filas = None
    try:
        pagina = libro[hoja] if hoja else libro.worksheets[0]
        filas = pagina.iter_rows(values_only=True)

        try:
            encabezados = [str(c).strip() if c is not None else "" for c in next(filas)]
        except StopIteration:
            return [], []

        for requerida in (columna_nombre, columna_telefono):
            if requerida not in encabezados:
                raise ValueError(
                    f"El Excel no tiene la columna '{requerida}'. "
                    f"Columnas encontradas: {encabezados}"
                )

        indice_nombre = encabezados.index(columna_nombre)
        indice_telefono = encabezados.index(columna_telefono)

        validos = []
        descartes = []
        ya_vistos = set()

        for numero_fila, fila in enumerate(filas, start=2):
            if not any(c is not None and str(c).strip() for c in fila):
                continue  # fila completamente vacia: ni siquiera es un descarte

            nombre_crudo = fila[indice_nombre] if indice_nombre < len(fila) else None
            telefono_crudo = fila[indice_telefono] if indice_telefono < len(fila) else None

            telefono, motivo = normalizar_telefono(telefono_crudo)
            if telefono is None:
                descartes.append(Descarte(
                    fila=numero_fila,
                    nombre_crudo=_texto(nombre_crudo),
                    telefono_crudo=_texto(telefono_crudo),
                    motivo=motivo,
                ))
                continue

            if telefono in ya_vistos:
                descartes.append(Descarte(
                    fila=numero_fila,
                    nombre_crudo=_texto(nombre_crudo),
                    telefono_crudo=_texto(telefono_crudo),
                    motivo=DUPLICADO,
                ))
                continue

            ya_vistos.add(telefono)
            validos.append(Contacto(
                nombre=normalizar_nombre(nombre_crudo, nombre_por_defecto),
                telefono=telefono,
                fila=numero_fila,
                crudo={
                    encabezado: fila[i] if i < len(fila) else None
                    for i, encabezado in enumerate(encabezados)
                    if encabezado
                },
            ))

        return validos, descartes
    finally:
        # En modo read_only, openpyxl abre el XML de la hoja como un stream
        # aparte. Si salimos antes de agotar el generador -por ejemplo al
        # detectar que falta una columna- ese stream sigue abierto aunque
        # cerremos el libro, y en Windows el archivo queda bloqueado: no se
        # puede borrar ni reemplazar.
        if filas is not None:
            filas.close()
        libro.close()


def _texto(valor):
    return "" if valor is None else str(valor).strip()
