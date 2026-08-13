"""Marca el Excel original con el estado de cada contacto.

Ojo: el CSV de `registro.py` sigue siendo la fuente de verdad. Esto es una
capa de visualizacion encima. Se escribe al FINAL de la corrida, nunca mensaje
por mensaje, porque guardar un .xlsx reescribe el archivo completo y un corte
a medias lo dejaria corrupto.
"""

import shutil
from datetime import datetime

from openpyxl import load_workbook

from src.contactos import normalizar_telefono

COLUMNA_ESTADO = "Estado"
COLUMNA_FECHA = "Fecha envio"
COLUMNA_ERROR = "Error"

COLUMNAS_NUEVAS = (COLUMNA_ESTADO, COLUMNA_FECHA, COLUMNA_ERROR)

PENDIENTE = "pendiente"
DESCARTADO = "descartado"


class ExcelBloqueado(Exception):
    """No se pudo guardar el Excel, casi siempre porque esta abierto."""


def marcar_excel(ruta, estados_por_telefono, hoja=None, columna_telefono="Teléfono"):
    """Escribe Estado / Fecha envio / Error en el Excel. Devuelve filas marcadas.

    `estados_por_telefono` viene del registro: {telefono: (estado, timestamp, error)}.
    Un contacto que no aparece ahi queda "pendiente"; uno cuyo telefono no es
    valido queda "descartado" con el motivo.

    Antes de guardar deja una copia en `<ruta>.bak`.
    """
    libro = load_workbook(ruta)
    try:
        pagina = libro[hoja] if hoja else libro.worksheets[0]

        encabezados = [
            str(c.value).strip() if c.value is not None else ""
            for c in pagina[1]
        ]

        if columna_telefono not in encabezados:
            raise ValueError(
                f"El Excel no tiene la columna '{columna_telefono}'. "
                f"Columnas encontradas: {encabezados}"
            )

        indice_telefono = encabezados.index(columna_telefono)
        columnas = _asegurar_columnas(pagina, encabezados)

        marcadas = 0
        for fila in pagina.iter_rows(min_row=2):
            if not any(c.value is not None and str(c.value).strip() for c in fila):
                continue  # fila vacia: no se marca

            crudo = fila[indice_telefono].value if indice_telefono < len(fila) else None
            telefono, motivo = normalizar_telefono(crudo)

            if telefono is None:
                estado, momento, error = DESCARTADO, "", motivo
            else:
                estado, momento, error = estados_por_telefono.get(
                    telefono, (PENDIENTE, "", "")
                )

            pagina.cell(row=fila[0].row, column=columnas[COLUMNA_ESTADO], value=estado)
            pagina.cell(row=fila[0].row, column=columnas[COLUMNA_FECHA],
                        value=_formatear_fecha(momento))
            pagina.cell(row=fila[0].row, column=columnas[COLUMNA_ERROR], value=error)
            marcadas += 1

        _guardar(libro, ruta)
        return marcadas
    finally:
        libro.close()


def _asegurar_columnas(pagina, encabezados):
    """Devuelve {nombre_columna: numero_de_columna}, creandolas si no existen.

    Si el Excel ya fue marcado antes, reusa las columnas en vez de duplicarlas.
    """
    columnas = {}
    siguiente = len(encabezados) + 1

    for nombre in COLUMNAS_NUEVAS:
        if nombre in encabezados:
            columnas[nombre] = encabezados.index(nombre) + 1
        else:
            pagina.cell(row=1, column=siguiente, value=nombre)
            columnas[nombre] = siguiente
            siguiente += 1

    return columnas


def _formatear_fecha(momento_iso):
    """Convierte el timestamp UTC del log a hora local legible."""
    if not momento_iso:
        return ""
    try:
        momento = datetime.fromisoformat(momento_iso)
    except ValueError:
        return momento_iso
    if momento.tzinfo is not None:
        momento = momento.astimezone()
    return momento.strftime("%Y-%m-%d %H:%M")


def _guardar(libro, ruta):
    try:
        shutil.copy2(ruta, ruta + ".bak")
    except OSError:
        pass  # sin copia de seguridad se puede seguir; no es motivo para abortar

    try:
        libro.save(ruta)
    except PermissionError:
        raise ExcelBloqueado(
            f"No pude escribir en {ruta}: el archivo esta abierto en Excel o WPS. "
            "Cierralo y vuelve a intentar. Los envios ya quedaron guardados en el log."
        )
