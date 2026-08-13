import os

import pytest
from openpyxl import Workbook, load_workbook

from src.marcado import (
    COLUMNA_ERROR,
    COLUMNA_ESTADO,
    COLUMNA_FECHA,
    ExcelBloqueado,
    marcar_excel,
)


def _crear_excel(tmp_path, filas, encabezados=("Nombre", "Email", "Teléfono")):
    wb = Workbook()
    ws = wb.active
    ws.append(list(encabezados))
    for fila in filas:
        ws.append(list(fila))
    ruta = tmp_path / "contactos.xlsx"
    wb.save(ruta)
    return str(ruta)


def _leer(ruta):
    """Devuelve (encabezados, filas) del Excel ya marcado."""
    wb = load_workbook(ruta)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    return list(filas[0]), filas[1:]


def test_agrega_las_tres_columnas_de_estado(tmp_path):
    ruta = _crear_excel(tmp_path, [("Ana", "a@x.com", 3001234567)])

    marcar_excel(ruta, {})

    encabezados, _ = _leer(ruta)
    assert encabezados[-3:] == [COLUMNA_ESTADO, COLUMNA_FECHA, COLUMNA_ERROR]


def test_marca_como_enviado_con_su_fecha(tmp_path):
    ruta = _crear_excel(tmp_path, [("Ana", "a@x.com", 3001234567)])
    estados = {"573001234567": ("enviado", "2026-08-13T14:54:03+00:00", "")}

    marcar_excel(ruta, estados)

    _, filas = _leer(ruta)
    assert filas[0][-3] == "enviado"
    assert filas[0][-2].startswith("2026-08-13")
    assert filas[0][-1] in (None, "")


def test_marca_como_fallo_con_el_motivo(tmp_path):
    ruta = _crear_excel(tmp_path, [("Ana", "a@x.com", 3001234567)])
    estados = {"573001234567": ("fallo", "2026-08-13T14:54:03+00:00", "131026 Recipient not found")}

    marcar_excel(ruta, estados)

    _, filas = _leer(ruta)
    assert filas[0][-3] == "fallo"
    assert "131026" in filas[0][-1]


def test_quien_no_esta_en_el_registro_queda_pendiente(tmp_path):
    ruta = _crear_excel(tmp_path, [("Ana", "a@x.com", 3001234567)])

    marcar_excel(ruta, {})

    _, filas = _leer(ruta)
    assert filas[0][-3] == "pendiente"


def test_un_telefono_invalido_queda_descartado_con_su_motivo(tmp_path):
    ruta = _crear_excel(tmp_path, [("Pedro", "p@x.com", "604 4488388")])

    marcar_excel(ruta, {})

    _, filas = _leer(ruta)
    assert filas[0][-3] == "descartado"
    assert filas[0][-1] == "no_es_movil_co"


def test_no_toca_las_columnas_originales(tmp_path):
    ruta = _crear_excel(tmp_path, [("Ana", "a@x.com", 3001234567)])

    marcar_excel(ruta, {"573001234567": ("enviado", "2026-08-13T14:54:03+00:00", "")})

    encabezados, filas = _leer(ruta)
    assert encabezados[:3] == ["Nombre", "Email", "Teléfono"]
    assert filas[0][:3] == ("Ana", "a@x.com", 3001234567)


def test_marcar_dos_veces_no_duplica_las_columnas(tmp_path):
    ruta = _crear_excel(tmp_path, [("Ana", "a@x.com", 3001234567)])

    marcar_excel(ruta, {})
    marcar_excel(ruta, {"573001234567": ("enviado", "2026-08-13T14:54:03+00:00", "")})

    encabezados, filas = _leer(ruta)
    assert encabezados.count(COLUMNA_ESTADO) == 1
    assert len(encabezados) == 6
    assert filas[0][-3] == "enviado"          # y si actualiza el valor


def test_devuelve_cuantas_filas_marco(tmp_path):
    ruta = _crear_excel(tmp_path, [
        ("Ana", "a@x.com", 3001234567),
        ("Luis", "l@x.com", 3009876543),
    ])

    assert marcar_excel(ruta, {}) == 2


def test_deja_una_copia_de_seguridad(tmp_path):
    ruta = _crear_excel(tmp_path, [("Ana", "a@x.com", 3001234567)])

    marcar_excel(ruta, {})

    assert os.path.exists(ruta + ".bak")


def test_avisa_claro_si_el_archivo_esta_bloqueado(tmp_path, monkeypatch):
    ruta = _crear_excel(tmp_path, [("Ana", "a@x.com", 3001234567)])

    def guardar_bloqueado(self, destino):
        raise PermissionError(13, "Permission denied")

    monkeypatch.setattr("openpyxl.workbook.workbook.Workbook.save", guardar_bloqueado)

    with pytest.raises(ExcelBloqueado) as error:
        marcar_excel(ruta, {})

    assert "abierto" in str(error.value).lower()


def test_ignora_filas_vacias(tmp_path):
    ruta = _crear_excel(tmp_path, [("Ana", "a@x.com", 3001234567), (None, None, None)])

    assert marcar_excel(ruta, {}) == 1

    # Solo la fila con datos queda marcada. openpyxl descarta al guardar las
    # filas que quedaron completamente vacias.
    _, filas = _leer(ruta)
    marcadas = [f for f in filas if f[-3] not in (None, "")]
    assert len(marcadas) == 1
    assert marcadas[0][0] == "Ana"
